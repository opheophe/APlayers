import re
import os
import sys
import csv
import json
import time
import glob
import threading
import configparser
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from tqdm import tqdm

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DELAY_MS = 2000
RETRIES = 10
PLAYER_AGE_CUTOFF = 18
REFETCH_PLAYERS = False
SOUND = True
EXPORT_PLAYERS = True
EVENT_LABELS = {
    "sb-aus": "Subbed out",
    "sb-ein": "Subbed in",
    "sb-gelb": "Yellow",
    "sb-gelb-rot": "2nd yellow",
    "sb-rot": "Red card",
    "sb-tor": "Goal",
    "sb-verletzung": "Injury",
}
OUTPUT_DIR = SCRIPT_DIR
LOG_DIR = os.path.join(SCRIPT_DIR, "Logs")
MAX_FILES = 10
MAX_LOGS = 10
DATA_FILE = os.path.join(SCRIPT_DIR, "Data.json")
SETTINGS_PATH = os.path.join(SCRIPT_DIR, "Settings.ini")
LIGOR_SECTION = "Ligor"
LIGOR_PREFIX = "Liga."

DEFAULT_RESPONSIBLE = "Not assigned"

DEFAULT_LIGOR = [
    {"id": 1, "name": "Premium Liiga", "country": "Estonia", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/EST1/saison_id/2026", "active": True},
    {"id": 2, "name": "Virsliga", "country": "Latvia", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/LET1/saison_id/2026", "active": True},
    {"id": 3, "name": "Toplyga", "country": "Lithuania", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/LI1/saison_id/2026", "active": True},
    {"id": 4, "name": "Kategoria Superiore", "country": "Albania", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/ALB1/saison_id/2026", "active": True},
    {"id": 5, "name": "Superliga e Kosovës", "country": "Kosovo", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/KO1/saison_id/2026", "active": True},
    {"id": 6, "name": "Premijer Liga Bosne i Hercegovine", "country": "Bosnia and Herzegovina", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/BOS1/saison_id/2026", "active": True},
    {"id": 7, "name": "Prva Makedonska Fudbalska Liga", "country": "North Macedonia", "year": "2026", "responsible": "Alen",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/MAZ1/saison_id/2026", "active": True},
    {"id": 8, "name": "Premier Liga", "country": "Ukraine", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/UKR1/saison_id/2026", "active": True},
    {"id": 9, "name": "efbet Liga", "country": "Bulgaria", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/BU1/saison_id/2026", "active": True},
    {"id": 10, "name": "Niké Liga", "country": "Slovakia", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/SLO1/saison_id/2026", "active": True},
    {"id": 11, "name": "O'zbekiston Superligasi", "country": "Uzbekistan", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/UZ1/saison_id/2026", "active": True},
    {"id": 12, "name": "Meridianbet 1. CFL", "country": "Montenegro", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/MNE1/saison_id/2026", "active": True},
    {"id": 13, "name": "Super Liga", "country": "Moldova", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/MO1N/saison_id/2026", "active": True},
    {"id": 14, "name": "SuperLiga", "country": "Romania", "year": "2026", "responsible": "Uwe",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/RO1/saison_id/2026", "active": True},
    {"id": 15, "name": "J1 League", "country": "Japan", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/JAP1/saison_id/2026", "active": True},
    {"id": 16, "name": "K League 1", "country": "South Korea", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/RSK1/saison_id/2026", "active": True},
    {"id": 17, "name": "BGL Ligue", "country": "Luxembourg", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/LUX1/saison_id/2026", "active": True},
    {"id": 18, "name": "Nemzeti Bajnokság", "country": "Hungary", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/UNG1/saison_id/2026", "active": True},
    {"id": 19, "name": "Super League 1", "country": "Greece", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/GR1/saison_id/2026", "active": True},
    {"id": 20, "name": "Süper Lig", "country": "Turkey", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/TR1/saison_id/2026", "active": True},
    {"id": 21, "name": "Cyprus League", "country": "Cyprus", "year": "2026", "responsible": "Dan",
     "url": "https://www.transfermarkt.com/toplyga/gesamtspielplan/wettbewerb/ZYP1/saison_id/2026", "active": True},
]


def load_settings():
    global DELAY_MS, RETRIES, EVENT_LABELS, PLAYER_AGE_CUTOFF, REFETCH_PLAYERS, SOUND, EXPORT_PLAYERS
    if not os.path.exists(SETTINGS_PATH):
        save_settings()
        return
    cfg = configparser.ConfigParser()
    cfg.read(SETTINGS_PATH)
    try:
        DELAY_MS = cfg.getint("Settings", "delay_ms", fallback=2000)
        RETRIES = cfg.getint("Settings", "retries", fallback=10)
        PLAYER_AGE_CUTOFF = cfg.getint("Settings", "player_age_cutoff", fallback=18)
        REFETCH_PLAYERS = cfg.getboolean("Settings", "refetch_players", fallback=False)
        SOUND = cfg.getboolean("Settings", "sound", fallback=True)
        EXPORT_PLAYERS = cfg.getboolean("Settings", "export_players", fallback=True)
    except Exception:
        pass
    if cfg.has_section("EventLabels"):
        for key, val in cfg.items("EventLabels"):
            EVENT_LABELS[key] = val


def save_settings():
    cfg = _load_config()
    cfg["Settings"] = {
        "delay_ms": str(DELAY_MS),
        "retries": str(RETRIES),
        "player_age_cutoff": str(PLAYER_AGE_CUTOFF),
        "refetch_players": str(REFETCH_PLAYERS),
        "sound": str(SOUND),
        "export_players": str(EXPORT_PLAYERS),
    }
    cfg["EventLabels"] = EVENT_LABELS
    _write_config(cfg)


def _load_config():
    cfg = configparser.ConfigParser()
    if os.path.exists(SETTINGS_PATH):
        cfg.read(SETTINGS_PATH, encoding="utf-8")
    return cfg


def _write_config(cfg):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        cfg.write(f)


def load_ligor():
    cfg = _load_config()
    ligor = []
    for section in cfg.sections():
        if not section.startswith(LIGOR_PREFIX):
            continue
        liga = {
            "id": _to_int(cfg.get(section, "id", fallback=section[len(LIGOR_PREFIX):])),
            "name": cfg.get(section, "name", fallback=""),
            "country": cfg.get(section, "country", fallback=""),
            "year": cfg.get(section, "year", fallback=""),
            "responsible": cfg.get(section, "responsible", fallback=DEFAULT_RESPONSIBLE),
            "url": cfg.get(section, "url", fallback=""),
            "active": cfg.getboolean(section, "active", fallback=True),
        }
        ligor.append(liga)
    if not ligor:
        ligor = [dict(l) for l in DEFAULT_LIGOR]
        save_ligor(ligor)
    ligor.sort(key=lambda l: l["id"])
    return ligor


def save_ligor(ligor):
    cfg = _load_config()
    for section in list(cfg.sections()):
        if section.startswith(LIGOR_PREFIX):
            cfg.remove_section(section)
    if LIGOR_SECTION in cfg.sections():
        cfg.remove_section(LIGOR_SECTION)

    ids = [l.get("id", 0) for l in ligor]
    next_id = max(ids) + 1 if ids else 1
    cfg[LIGOR_SECTION] = {"next_id": str(next_id)}

    for liga in ligor:
        section = f"{LIGOR_PREFIX}{liga.get('id', 0)}"
        cfg[section] = {
            "id": str(liga.get("id", 0)),
            "name": str(liga.get("name", "")),
            "country": str(liga.get("country", "")),
            "year": str(liga.get("year", "")),
            "responsible": str(liga.get("responsible", DEFAULT_RESPONSIBLE)),
            "url": str(liga.get("url", "")),
            "active": str(bool(liga.get("active", True))),
        }
    _write_config(cfg)


def next_liga_id(ligor):
    ids = [l.get("id", 0) for l in ligor]
    return max(ids) + 1 if ids else 1


def _to_int(value):
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return 0


FILTER_KEYS = ("responsible", "liga", "year", "country", "section", "position", "age",
               "player_age", "player_country", "player_league", "player_year")

PLAYER_FILTER_KEYS = ("player_age", "player_country", "player_league", "player_year")


def load_filters():
    cfg = _load_config()
    filters = {key: [] for key in FILTER_KEYS}
    if cfg.has_section("Filters"):
        for key in FILTER_KEYS:
            if cfg.has_option("Filters", key):
                raw = cfg.get("Filters", key)
                try:
                    val = json.loads(raw)
                    if isinstance(val, list):
                        filters[key] = val
                    elif val:
                        filters[key] = [val]
                except ValueError:
                    filters[key] = [raw] if raw and raw != "All" else []
    return filters


def save_filters(filters):
    cfg = _load_config()
    cfg["Filters"] = {key: json.dumps(filters.get(key) or []) for key in FILTER_KEYS}
    _write_config(cfg)


def load_columns():
    cfg = _load_config()
    cols = {}
    for i, key in enumerate(COLUMN_KEYS):
        d = DEFAULT_COLUMNS.get(key, {"lista": True, "excel": True})
        cols[key] = {"lista": d["lista"], "excel": d["excel"], "order": (i + 1) * 10}
    if cfg.has_section("Columns"):
        for key in COLUMN_KEYS:
            if cfg.has_option("Columns", key + "_lista"):
                cols[key]["lista"] = cfg.getboolean("Columns", key + "_lista", fallback=True)
            if cfg.has_option("Columns", key + "_excel"):
                cols[key]["excel"] = cfg.getboolean("Columns", key + "_excel", fallback=True)
            if cfg.has_option("Columns", key + "_order"):
                cols[key]["order"] = cfg.getint("Columns", key + "_order", fallback=(COLUMN_KEYS.index(key) + 1) * 10)
    return cols


def save_columns(cols):
    cfg = _load_config()
    data = {}
    for key in COLUMN_KEYS:
        c = cols.get(key, {"lista": True, "excel": True})
        data[key + "_lista"] = str(bool(c.get("lista", True)))
        data[key + "_excel"] = str(bool(c.get("excel", True)))
        data[key + "_order"] = str(int(c.get("order", (COLUMN_KEYS.index(key) + 1) * 10)))
    cfg["Columns"] = data
    _write_config(cfg)


def visible_columns(cols, flag):
    def key(k):
        c = cols.get(k, {})
        order = c.get("order")
        if order is None:
            order = (COLUMN_KEYS.index(k) + 1) * 10
        return (order, COLUMN_LABELS.get(k, k))

    return sorted([k for k in COLUMN_KEYS if cols.get(k, {}).get(flag, True)], key=key)


def load_window_size():
    cfg = _load_config()
    width = 900
    height = 650
    if cfg.has_section("Window"):
        width = cfg.getint("Window", "width", fallback=900)
        height = cfg.getint("Window", "height", fallback=650)
    return width, height


def save_window_size(width, height):
    cfg = _load_config()
    cfg["Window"] = {"width": str(int(width)), "height": str(int(height))}
    _write_config(cfg)


abort_check = lambda: False
_match_urls = {}

log_handle = None

try:
    os.system("")
except Exception:
    pass
C = {
    "R": "\033[91m",
    "G": "\033[92m",
    "Y": "\033[93m",
    "B": "\033[94m",
    "M": "\033[95m",
    "C": "\033[96m",
    "W": "\033[97m",
    "BR": "\033[1;91m",
    "BG": "\033[1;92m",
    "BY": "\033[1;93m",
    "BB": "\033[1;94m",
    "BC": "\033[1;96m",
    "BW": "\033[1;97m",
    "X": "\033[0m",
}


def log(msg, color=None):
    is_tty = sys.stdout is not None and sys.stdout.isatty()
    text = f"{C.get(color, '')}{msg}{C['X']}" if color and is_tty else msg
    if sys.stdout:
        tqdm.write(text)
    if log_handle:
        log_handle.write(msg + "\n")
        log_handle.flush()


def fetch(url, timeout=30, retries=None):
    if retries is None:
        retries = RETRIES

    def interruptible_sleep(seconds):
        if seconds <= 0:
            return True
        end = time.time() + seconds
        while time.time() < end:
            if abort_check():
                return False
            time.sleep(min(0.2, end - time.time()))
        return True

    def do_request():
        result = [None]
        err = [None]
        def worker():
            try:
                result[0] = requests.get(url, headers=HEADERS, timeout=timeout)
            except Exception as e:
                err[0] = e
        t = threading.Thread(target=worker, daemon=True)
        t.start()
        while t.is_alive():
            if abort_check():
                return None
            t.join(0.2)
        if err[0]:
            raise err[0]
        return result[0]

    for attempt in range(retries):
        if abort_check():
            return None
        try:
            if attempt > 0:
                wait = 2 ** attempt
                log(f"  Retry {attempt} after {wait}s...", "Y")
                if not interruptible_sleep(wait):
                    return None
            else:
                if not interruptible_sleep(DELAY_MS / 1000):
                    return None
            resp = do_request()
            if resp is None:
                return None
            if resp.status_code in (429, 503):
                log(f"  Server returned {resp.status_code}, retrying...", "Y")
                if not interruptible_sleep(5 * (attempt + 1)):
                    return None
                continue
            return resp
        except requests.exceptions.RequestException as e:
            log(f"  Request failed (attempt {attempt + 1}/{retries}): {e}", "Y")
    return None


def fetch_league_info(url):
    resp = fetch(url)
    name = ""
    country = ""
    if resp is not None and resp.status_code == 200:
        soup = BeautifulSoup(resp.text, "html.parser")
        keywords = soup.find("meta", attrs={"name": "keywords"})
        if keywords and keywords.get("content"):
            parts = [p.strip() for p in keywords["content"].split(",")]
            if parts:
                name = parts[0]
            if len(parts) > 1:
                country = parts[1]
        if not name:
            og_title = soup.find("meta", attrs={"property": "og:title"})
            if og_title and og_title.get("content"):
                name = og_title["content"].split(" - ")[0].strip()
    m = re.search(r"saison_id/(\d{4})", url or "")
    year = m.group(1) if m else ""
    return name, country, year


def extract_result_links(name, url):
    _match_urls[name] = url
    log(f"[{name}] Hämtar matcher...")
    resp = fetch(url)
    if resp is None or resp.status_code != 200:
        log(f"[{name}] Misslyckades")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    links = []

    for td in soup.find_all("td", class_="zentriert"):
        text = td.get_text(strip=True)
        if re.match(r"^\d+:\d+$", text):
            a = td.find("a", href=True)
            if a and "spielbericht" in a["href"]:
                links.append("https://www.transfermarkt.com" + a["href"])

    return links


def parse_match(country, match_url, compact=False):
    lineup_url = match_url.replace("/index/", "/aufstellung/")
    match_id = match_url.rstrip("/").rsplit("/", 1)[-1]
    _match_urls[match_id] = lineup_url
    log(f"  Hämtar lineup: {country} {match_id}")
    resp = fetch(lineup_url)
    if resp is None or resp.status_code != 200:
        log("  MISSLYCKADES efter alla försök", "BR")
        return [], None

    soup = BeautifulSoup(resp.text, "html.parser")

    box = soup.select_one(".box.sb-spielbericht-head")
    if not box:
        return [], None

    liga_tag = box.select_one(".direct-headline__header a")
    liga = liga_tag.get_text(strip=True) if liga_tag else "?"

    matchday_tag = box.select_one(".sb-datum a")
    matchday = matchday_tag.get_text(strip=True) if matchday_tag else "?"

    datum_tag = box.select_one(".sb-datum a[href*='waspassiertheute']")
    datum_raw = datum_tag.get_text(strip=True) if datum_tag else ""
    datum = datum_raw
    if datum_raw:
        try:
            dt = datetime.strptime(datum_raw, "%a, %d/%m/%y")
            datum = dt.strftime("%Y-%m-%d")
        except ValueError:
            pass

    zeit_p = box.select_one(".sb-datum")
    zeit = ""
    if zeit_p:
        full_text = zeit_p.get_text(" ", strip=True)
        parts = full_text.split("|")
        zeit = parts[-1].strip().replace("\xa0", " ").replace("��", "") if len(parts) >= 3 else ""

    ergebnis = box.select_one(".sb-endstand")
    ergebnis_text = ergebnis.get_text(" ", strip=True).replace("  ", " ") if ergebnis else "?"

    heim = box.select_one(".sb-heim a.sb-vereinslink")
    heim_name = heim.get_text(strip=True) if heim else "?"

    gast = box.select_one(".sb-gast a.sb-vereinslink")
    gast_name = gast.get_text(strip=True) if gast else "?"

    meta = {
        "liga": liga,
        "home": heim_name,
        "away": gast_name,
        "matchday": matchday,
        "date": datum,
        "time": zeit,
        "result": ergebnis_text,
    }
    if compact:
        meta["country"] = country

    rows = []
    team_names = [heim_name, gast_name]
    team_toggle = 0

    for heading in soup.find_all("h2"):
        section = heading.get_text(strip=True)
        if section not in ("Starting Line-up", "Substitutes", "Manager"):
            continue

        container = heading.find_parent("div", class_="large-6")
        if not container:
            continue

        team = team_names[team_toggle]
        team_toggle = 1 - team_toggle

        table = container.select_one(".responsive-table > table")
        if not table:
            continue

        for tr in table.find_all("tr", recursive=False):
            if section == "Manager":
                parsed = parse_manager_row(tr)
            else:
                parsed = parse_player_row(tr)

            if parsed:
                number, name, player_id, role, salary, age, nationality, events = parsed
                if compact:
                    rows.append([section, team, number, name, player_id, role, salary, age, nationality, events])
                else:
                    fixture = f"{heim_name} vs {gast_name}" if heim_name and gast_name else ""
                    rows.append(
                        [country, liga, matchday, team, datum, zeit, ergebnis_text, section, number, name, player_id, role, salary, age, nationality, fixture, match_url, events]
                    )

    return rows, meta


def parse_player_row(row):
    num_div = row.select_one(".rn_nummer")
    number = num_div.get_text(strip=True) if num_div else "-"

    name_tag = row.select_one("a.wichtig")
    name = name_tag.get_text(strip=True) if name_tag else "?"
    player_id = ""
    if name_tag:
        href = name_tag.get("href", "")
        m = re.search(r"/spieler/(\d+)", href)
        if m:
            player_id = m.group(1)

    cells = row.find_all("td", recursive=False)
    if len(cells) < 3:
        return None

    inline_table = cells[1].find("table", class_="inline-table")
    if not inline_table:
        return None

    rows = inline_table.find_all("tr")
    if len(rows) < 2:
        return None

    text_row1 = rows[0].get_text(" ", strip=True)
    text_row2 = rows[1].get_text(strip=True)

    age_match = re.search(r"\((\d+) years old\)", text_row1)
    age = age_match.group(1) if age_match else ""

    role = ""
    salary = ""
    if "," in text_row2:
        role, salary = text_row2.split(",", 1)
        role = role.strip()
        salary = salary.strip()
    else:
        role = text_row2.strip()

    flag = cells[2].find("img", class_="flaggenrahmen") if len(cells) > 2 else None
    nationality = flag.get("title", "") if flag else ""

    event_spans = rows[0].find_all("span", class_=lambda c: c and "sb-sprite" in c)
    events = " ".join(s.get("class")[-1] for s in event_spans)

    return [number, name, player_id, role, salary, age, nationality, events]


def parse_manager_row(row):
    name_tag = row.select_one("a.wichtig")
    name = name_tag.get_text(strip=True) if name_tag else "?"
    player_id = ""
    if name_tag:
        href = name_tag.get("href", "")
        m = re.search(r"/spieler/(\d+)", href)
        if m:
            player_id = m.group(1)

    cells = row.find_all("td", recursive=False)
    if len(cells) < 2:
        return None

    inline_table = cells[0].find("table", class_="inline-table")
    if not inline_table:
        return None

    rows = inline_table.find_all("tr")
    if len(rows) < 2:
        return None

    age_text = rows[1].get_text(strip=True)
    age_match = re.search(r"(\d+) years old", age_text)
    age = age_match.group(1) if age_match else ""

    flag = cells[1].find("img", class_="flaggenrahmen") if len(cells) > 1 else None
    nationality = flag.get("title", "") if flag else ""

    event_spans = rows[0].find_all("span", class_=lambda c: c and "sb-sprite" in c)
    events = " ".join(s.get("class")[-1] for s in event_spans)

    return ["-", name, player_id, "Manager", "", age, nationality, events]


def rotate_files(pattern, max_count):
    files = sorted(glob.glob(pattern))
    while len(files) >= max_count:
        oldest = files.pop(0)
        os.remove(oldest)
        log(f"  Removed old file: {oldest}")


def process_pending_games(data, active_ids=None, progress_cb=None):
    success = 0
    total = 0
    for lid, league in data.get("Leagues", {}).items():
        if active_ids is not None and league.get("Id") not in active_ids:
            continue
        country = league.get("Country", "?")
        for gid, game in league.get("Games", {}).items():
            if game.get("Status") != "pending":
                continue
            total += 1
            url = game.get("Url", "")
            try:
                rows, meta = parse_match(country, url, compact=True)
            except Exception as e:
                log(f"  Skipped {url}: {e}", "Y")
                continue

            if meta is None:
                continue

            if rows:
                game["Status"] = "success"
                game["Meta"] = meta
                game["Lineup"] = rows
                add_players_from_lineup(data, rows)
                success += 1

            save_data(data)

    return success


def load_data():
    if not os.path.exists(DATA_FILE):
        return {"Leagues": {}, "Players": {}}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "Leagues" not in data:
        data["Leagues"] = {}
    if "Players" not in data:
        data["Players"] = {}
    return data


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def ensure_league(data, liga):
    lid = str(liga.get("id"))
    leagues = data.setdefault("Leagues", {})
    league = leagues.get(lid)
    if league is None:
        league = {"Id": liga.get("id"), "Games": {}}
        leagues[lid] = league
    league["Name"] = liga.get("name", "")
    league["Country"] = liga.get("country", "")
    league["Year"] = liga.get("year", "")
    league["Responsible"] = liga.get("responsible", DEFAULT_RESPONSIBLE)
    league["Url"] = liga.get("url", "")
    league.setdefault("Games", {})
    return league


def match_filter_league(league, active_ids, sel):
    if active_ids is not None and league.get("Id") not in active_ids:
        return False
    if not sel:
        return True
    if sel.get("responsible") and league.get("Responsible") not in sel["responsible"]:
        return False
    if sel.get("liga") and league.get("Name") not in sel["liga"]:
        return False
    if sel.get("year") and league.get("Year") not in sel["year"]:
        return False
    if sel.get("country") and league.get("Country") not in sel["country"]:
        return False
    return True


def filter_ligor(ligor, sel):
    result = []
    for l in ligor:
        if sel.get("responsible") and l.get("responsible") not in sel["responsible"]:
            continue
        if sel.get("liga") and l.get("name") not in sel["liga"]:
            continue
        if sel.get("year") and l.get("year") not in sel["year"]:
            continue
        if sel.get("country") and l.get("country") not in sel["country"]:
            continue
        result.append(l)
    return result


def match_filter_player(row, sel):
    if not sel:
        return True
    if sel.get("section") and (len(row) < 1 or row[0] not in sel["section"]):
        return False
    if sel.get("position") and (len(row) < 6 or row[5] not in sel["position"]):
        return False
    if sel.get("age") and (len(row) < 8 or row[7] not in sel["age"]):
        return False
    return True


def add_players_from_lineup(data, rows):
    players = data.setdefault("Players", {})
    added = 0
    for p in rows:
        if len(p) < 10:
            continue
        name = p[3]
        player_id = p[4]
        age = p[7]
        if not player_id or player_id in players:
            continue
        players[player_id] = {"Name": name, "Age": age}
        added += 1
    return added


def collect_players(data):
    added = 0
    for lid, league in data.get("Leagues", {}).items():
        for gid, game in league.get("Games", {}).items():
            if game.get("Status") != "success":
                continue
            added += add_players_from_lineup(data, game.get("Lineup", []))
    return added


def _slugify(name):
    slug = re.sub(r"[^\w\s-]", "", name.lower())
    slug = re.sub(r"\s+", "-", slug).strip("-")
    return slug or "player"


def player_profile_url(name, player_id):
    return f"https://www.transfermarkt.com/{_slugify(name)}/profil/spieler/{player_id}"


def fetch_player_profile(name, player_id):
    resp = fetch(player_profile_url(name, player_id))
    if resp is None or resp.status_code != 200:
        return None
    soup = BeautifulSoup(resp.text, "html.parser")
    info = {}

    club_a = soup.select_one("span.data-header__club a")
    if club_a:
        info["club"] = club_a.get_text(strip=True) or club_a.get("title", "")

    dob = soup.select_one("span[itemprop='birthDate']")
    if dob:
        info["dob"] = dob.get_text(strip=True).split("(")[0].strip()

    nat = soup.select_one("span[itemprop='nationality']")
    if nat:
        info["country"] = nat.get_text(" ", strip=True)

    height = soup.select_one("span[itemprop='height']")
    if height:
        info["height"] = height.get_text(strip=True)

    for li in soup.select("li.data-header__label"):
        label = li.get_text(" ", strip=True)
        if label.startswith("Position:"):
            content = li.select_one("span.data-header__content")
            if content:
                info["position"] = content.get_text(strip=True)
            break

    return info


def fetch_player_performance(player_id):
    url = f"https://www.transfermarkt.com/ceapi/performance-game/{player_id}"
    resp = fetch(url)
    if resp is None or resp.status_code != 200:
        return None
    try:
        data = resp.json()
    except Exception:
        return None
    if not data.get("success"):
        return None
    return (data.get("data") or {}).get("performance") or []


IDENTITY_FIELDS = {"shirtNumber", "positionId", "age", "primaryClubId", "injuryId",
                   "absenceId", "ageDiscrepancyDays", "grade", "pointsOnThePitch", "fairPlayPoints"}
RATIO_FIELDS = {"tacklesWonRatio", "passesReachedRatio", "scoringAttemptsOnGoalRatio"}
BOOL_FIELDS = {"isCaptain", "isStarting"}


def aggregate_performance(entries):
    groups = {}
    order = []
    for e in entries:
        gi = e.get("gameInformation") or {}
        stats = e.get("statistics") or {}
        league = gi.get("competitionId") or "?"
        season = gi.get("seasonId")
        key = (league, season)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(stats)

    result = []
    for league, season in order:
        sums = {}
        ratio_sum = {}
        ratio_cnt = {}
        last = {}
        bool_cnt = {}
        games = 0
        for stats in groups[(league, season)]:
            games += 1
            for cat, catval in stats.items():
                if not isinstance(catval, dict):
                    continue
                for k, v in catval.items():
                    if v is None:
                        continue
                    if k in BOOL_FIELDS:
                        bool_cnt[k] = bool_cnt.get(k, 0) + (1 if v else 0)
                    elif k in RATIO_FIELDS:
                        ratio_sum[k] = ratio_sum.get(k, 0.0) + v
                        ratio_cnt[k] = ratio_cnt.get(k, 0) + 1
                    elif k in IDENTITY_FIELDS:
                        last[k] = v
                    elif isinstance(v, (int, float)):
                        sums[k] = sums.get(k, 0) + v

        entry = {"league": league, "season": season, "games": games}
        for k, v in sums.items():
            entry[k] = v
        for k, v in ratio_sum.items():
            cnt = ratio_cnt.get(k, 0)
            entry[k] = round(v / cnt, 2) if cnt else None
        for k, v in bool_cnt.items():
            entry[k] = v
        for k, v in last.items():
            entry[k] = v
        result.append(entry)

    return result


DETAIL_FIELDS = ("Club", "Dob", "Position", "Country", "Height", "Performance", "Detailed")


def fetch_player_details(data, cutoff, progress_cb=None, refetch=False):
    players = data.get("Players", {})
    fetched = 0
    total = len(players)

    for i, (pid, player) in enumerate(players.items()):
        if abort_check():
            break

        age_raw = str(player.get("Age", "")).strip()
        age = int(age_raw) if age_raw.isdigit() else None
        if age is None or age > cutoff:
            continue
        if player.get("Detailed") and not refetch:
            continue

        name = player.get("Name", "")
        if not pid:
            continue

        if progress_cb:
            progress_cb(i + 1, total, name)

        for field in DETAIL_FIELDS:
            player.pop(field, None)

        profile = fetch_player_profile(name, pid)
        performance = fetch_player_performance(pid)

        if profile:
            player["Club"] = profile.get("club", "")
            player["Dob"] = profile.get("dob", "")
            player["Position"] = profile.get("position", "")
            player["Country"] = profile.get("country", "")
            player["Height"] = profile.get("height", "")
            player["Detailed"] = True
        if performance is not None:
            player["Performance"] = aggregate_performance(performance)

        fetched += 1
        save_data(data)

    return fetched


def format_duration(sec):
    m, s = divmod(int(sec), 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}h {m}m {s}s"
    elif m:
        return f"{m}m {s}s"
    return f"{s}s"


CSV_HEADER = ["responsible", "liga", "country", "year", "matchday", "fixture", "team", "date", "time", "result", "section", "player_id", "number", "name", "role", "salary", "age", "nationality", "url", "events"]

PERFORMANCE_COLUMNS = [
    "goalsScoredTotal", "assists", "scorer", "ownGoalsScored",
    "teamGoalsOnThePitch", "opponentGoalsOnThePitch",
    "penaltyShooterAttempts", "penaltyShooterGoalsScored",
    "penaltyShooterSaves", "penaltyShooterMisses",
    "penaltyGoalkeeperAttempts", "penaltyGoalkeeperGoalsConceded",
    "penaltyGoalkeeperSaves", "penaltyGoalkeeperMisses",
    "yellowCardNet", "yellowCardGross", "playedMinutes",
]

PLAYER_EXPORT_HEADER = ["id", "name", "age", "club", "dob", "country", "height", "position", "league", "year"] + PERFORMANCE_COLUMNS

PLAYER_COLUMN_KEYS = list(PLAYER_EXPORT_HEADER)

PLAYER_COLUMN_LABELS = {
    "id": "Id",
    "name": "Name",
    "age": "Age",
    "club": "Club",
    "dob": "Dob",
    "country": "Country",
    "height": "Height",
    "position": "Position",
    "league": "League",
    "year": "Year",
}
for _k in PERFORMANCE_COLUMNS:
    PLAYER_COLUMN_LABELS[_k] = _k

PLAYER_DEFAULT_COLUMNS = {key: {"lista": True, "excel": True} for key in PLAYER_COLUMN_KEYS}


def load_player_columns():
    cfg = _load_config()
    cols = {}
    for i, key in enumerate(PLAYER_COLUMN_KEYS):
        d = PLAYER_DEFAULT_COLUMNS.get(key, {"lista": True, "excel": True})
        cols[key] = {"lista": d["lista"], "excel": d["excel"], "order": (i + 1) * 10}
    if cfg.has_section("PlayerColumns"):
        for key in PLAYER_COLUMN_KEYS:
            if cfg.has_option("PlayerColumns", key + "_lista"):
                cols[key]["lista"] = cfg.getboolean("PlayerColumns", key + "_lista", fallback=True)
            if cfg.has_option("PlayerColumns", key + "_excel"):
                cols[key]["excel"] = cfg.getboolean("PlayerColumns", key + "_excel", fallback=True)
            if cfg.has_option("PlayerColumns", key + "_order"):
                cols[key]["order"] = cfg.getint("PlayerColumns", key + "_order", fallback=(PLAYER_COLUMN_KEYS.index(key) + 1) * 10)
    return cols


def save_player_columns(cols):
    cfg = _load_config()
    data = {}
    for key in PLAYER_COLUMN_KEYS:
        c = cols.get(key, {"lista": True, "excel": True})
        data[key + "_lista"] = str(bool(c.get("lista", True)))
        data[key + "_excel"] = str(bool(c.get("excel", True)))
        data[key + "_order"] = str(int(c.get("order", (PLAYER_COLUMN_KEYS.index(key) + 1) * 10)))
    cfg["PlayerColumns"] = data
    _write_config(cfg)


def visible_player_columns(cols, flag):
    def key(k):
        c = cols.get(k, {})
        order = c.get("order")
        if order is None:
            order = (PLAYER_COLUMN_KEYS.index(k) + 1) * 10
        return (order, PLAYER_COLUMN_LABELS.get(k, k))

    return sorted([k for k in PLAYER_COLUMN_KEYS if cols.get(k, {}).get(flag, True)], key=key)

COLUMN_KEYS = ["responsible", "liga", "country", "year", "matchday", "fixture", "team", "date", "time",
               "result", "section", "player_id", "number", "name", "role", "salary", "age",
               "nationality", "url", "events"]

COLUMN_LABELS = {
    "responsible": "Responsible",
    "liga": "Liga",
    "country": "Country",
    "year": "Year",
    "matchday": "Matchday",
    "team": "Team",
    "date": "Date",
    "time": "Time",
    "result": "Result",
    "section": "Section",
    "number": "Number",
    "name": "Name",
    "player_id": "Player ID",
    "role": "Role",
    "salary": "Salary",
    "age": "Age",
    "nationality": "Nationality",
    "fixture": "Fixture",
    "url": "URL",
    "events": "Events",
}

DEFAULT_COLUMNS = {
    "responsible": {"lista": False, "excel": True},
    "liga": {"lista": False, "excel": True},
    "country": {"lista": True, "excel": True},
    "year": {"lista": False, "excel": True},
    "matchday": {"lista": False, "excel": True},
    "fixture": {"lista": True, "excel": True},
    "team": {"lista": True, "excel": True},
    "date": {"lista": False, "excel": True},
    "time": {"lista": False, "excel": True},
    "result": {"lista": False, "excel": True},
    "section": {"lista": True, "excel": True},
    "player_id": {"lista": False, "excel": True},
    "number": {"lista": False, "excel": True},
    "name": {"lista": True, "excel": True},
    "role": {"lista": True, "excel": True},
    "salary": {"lista": False, "excel": True},
    "age": {"lista": True, "excel": True},
    "nationality": {"lista": False, "excel": True},
    "url": {"lista": True, "excel": True},
    "events": {"lista": True, "excel": True},
}


def translate_events(events_str):
    if not events_str:
        return ""
    parts = events_str.split()
    translated = [EVENT_LABELS.get(p, p) for p in parts]
    return " ".join(translated)


def _iter_export_rows(data, active_ids=None, sel=None):
    for lid, league in data.get("Leagues", {}).items():
        if not match_filter_league(league, active_ids, sel):
            continue
        responsible = league.get("Responsible", "")
        liga = league.get("Name", "?")
        country = league.get("Country", "?")
        year = league.get("Year", "")
        for gid, game in league.get("Games", {}).items():
            if game.get("Status") != "success":
                continue
            meta = game.get("Meta", {})
            matchday = meta.get("matchday", "?")
            date = meta.get("date", "?")
            time_val = meta.get("time", "?")
            result = meta.get("result", "?")
            home = meta.get("home", "")
            away = meta.get("away", "")
            fixture = f"{home} vs {away}" if home and away else ""
            match_url = game.get("Url", "")
            for p in game.get("Lineup", []):
                if not match_filter_player(p, sel):
                    continue
                if len(p) >= 10:
                    section, team, number, name, player_id, role, salary, age, nationality = p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7], p[8]
                    events = translate_events(p[9])
                else:
                    section, team, number, name, role, salary, age, nationality = p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7]
                    events = translate_events(p[8] if len(p) > 8 else "")
                    player_id = ""
                yield [responsible, liga, country, year, matchday, fixture, team, date, time_val, result, section, player_id, number, name, role, salary, age, nationality, match_url, events]


def iter_export_rows(data, active_ids=None, sel=None):
    yield from _iter_export_rows(data, active_ids, sel)


def _player_row(pid, p, entry):
    row = [pid, p.get("Name", ""), p.get("Age", ""), p.get("Club", ""),
           p.get("Dob", ""), p.get("Country", ""), p.get("Height", ""),
           p.get("Position", "")]
    if entry:
        row.append(entry.get("league", ""))
        row.append(entry.get("season", ""))
        for k in PERFORMANCE_COLUMNS:
            row.append(entry.get(k, ""))
    else:
        row.append("")
        row.append("")
        row.extend([""] * len(PERFORMANCE_COLUMNS))
    return row


def iter_player_export_rows(data, sel=None):
    sel = sel or {}
    for pid, p in data.get("Players", {}).items():
        age = str(p.get("Age", ""))
        country = p.get("Country", "")
        if sel.get("player_age") and age not in sel["player_age"]:
            continue
        if sel.get("player_country") and country not in sel["player_country"]:
            continue

        perf = p.get("Performance", [])
        if not perf:
            if sel.get("player_league") or sel.get("player_year"):
                continue
            yield _player_row(pid, p, None)
            continue

        for entry in perf:
            league = str(entry.get("league", ""))
            season = str(entry.get("season", ""))
            if sel.get("player_league") and league not in sel["player_league"]:
                continue
            if sel.get("player_year") and season not in sel["player_year"]:
                continue
            yield _player_row(pid, p, entry)


def export_csv(data, output_path, active_ids=None, sel=None):
    written = 0
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADER)
        for row in _iter_export_rows(data, active_ids, sel):
            writer.writerow(row)
            written += 1
    return written


def export_excel(data, output_path, active_ids=None, sel=None, columns=None, include_players=False, player_sel=None, player_columns=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl krävs för Excel-export. Installera med: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    header_border = Border(
        left=Side(style="thin", color="2F5496"),
        right=Side(style="thin", color="2F5496"),
        top=Side(style="thin", color="2F5496"),
        bottom=Side(style="thin", color="2F5496"),
    )

    TAB_COLORS = ["FF4444", "FF8C00", "FFD700", "44CC44", "4488FF", "CC44CC", "00CCCC", "FF69B4"]

    if columns is None:
        columns = COLUMN_KEYS
    header = [COLUMN_LABELS[k] for k in columns]
    col_indices = [COLUMN_KEYS.index(k) for k in columns]
    url_col = (columns.index("url") + 1) if "url" in columns else None

    matches_by_user = {}
    for row in _iter_export_rows(data, active_ids, sel):
        responsible = row[0]
        if responsible not in matches_by_user:
            matches_by_user[responsible] = []
        matches_by_user[responsible].append(row)

    total_written = 0
    sheet_count = 0

    for user in sorted(matches_by_user.keys()):
        safe_name = user[:31]
        ws = wb.create_sheet(title=safe_name)
        ws.sheet_properties.tabColor = TAB_COLORS[(sheet_count) % len(TAB_COLORS)]
        sheet_count += 1

        for col_idx, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        ws.freeze_panes = "A2"

        row_num = 2
        col_max = [len(h) for h in header]
        for full_row in matches_by_user[user]:
            values = [full_row[i] for i in col_indices]
            for col_idx, value in enumerate(values):
                cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                v_len = len(str(value)) if value is not None else 0
                if v_len > col_max[col_idx]:
                    col_max[col_idx] = v_len
            match_url = full_row[18] if len(full_row) > 18 else ""
            if match_url and url_col:
                ws.cell(row=row_num, column=url_col).hyperlink = match_url
            row_num += 1
            total_written += 1

        for col_idx, max_w in enumerate(col_max):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = min(max_w + 3, 150)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{row_num - 1}"

    if include_players:
        player_ws = wb.create_sheet(title="Spelare Detaljlista")
        player_ws.sheet_properties.tabColor = "7030A0"
        if player_columns is None:
            player_columns = PLAYER_COLUMN_KEYS
        player_header = [PLAYER_COLUMN_LABELS.get(k, k) for k in player_columns]
        player_col_indices = [PLAYER_COLUMN_KEYS.index(k) for k in player_columns]

        for col_idx, h in enumerate(player_header, 1):
            cell = player_ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        player_ws.freeze_panes = "A2"

        player_row_num = 2
        player_col_max = [len(h) for h in player_header]
        for prow in iter_player_export_rows(data, player_sel):
            values = [prow[i] for i in player_col_indices]
            for col_idx, value in enumerate(values):
                cell = player_ws.cell(row=player_row_num, column=col_idx + 1, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                v_len = len(str(value)) if value is not None else 0
                if v_len > player_col_max[col_idx]:
                    player_col_max[col_idx] = v_len
            player_row_num += 1

        for col_idx, max_w in enumerate(player_col_max):
            col_letter = get_column_letter(col_idx + 1)
            player_ws.column_dimensions[col_letter].width = min(max_w + 3, 150)

        player_ws.auto_filter.ref = f"A1:{get_column_letter(len(player_header))}{player_row_num - 1}"

    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.sheet_properties.tabColor = "548235"
    wb.move_sheet(summary_ws, offset=-(sheet_count + (1 if include_players else 0)))

    summary_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    summary_header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    for col_idx, h in enumerate(["Responsible", "Matches", "Players"], 1):
        cell = summary_ws.cell(row=1, column=col_idx, value=h)
        cell.font = summary_header_font
        cell.fill = summary_header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    summary_row = 2
    for user in sorted(matches_by_user.keys()):
        user_rows = matches_by_user[user]
        count = len(user_rows)
        summary_ws.cell(row=summary_row, column=1, value=user).border = thin_border
        summary_ws.cell(row=summary_row, column=2, value=count).border = thin_border
        summary_row += 1

    summary_ws.column_dimensions["A"].width = 18
    summary_ws.column_dimensions["B"].width = 10
    summary_ws.column_dimensions["C"].width = 12
    summary_ws.freeze_panes = "A2"

    wb.save(output_path)
    return total_written


def main():
    global log_handle

    t_start = datetime.now()
    now = t_start
    date_str = now.strftime("%Y-%m-%d")
    csv_path = os.path.join(OUTPUT_DIR, f"Output {date_str}.csv")

    os.makedirs(LOG_DIR, exist_ok=True)
    log_path = os.path.join(LOG_DIR, f"{date_str}.log")
    log_handle = open(log_path, "a", encoding="utf-8")

    log(f"=== Run started at {now.strftime('%Y-%m-%d %H:%M:%S')} ===", "C")

    rotate_files(os.path.join(OUTPUT_DIR, "Output *.csv"), MAX_FILES)
    rotate_files(os.path.join(LOG_DIR, "*.log"), MAX_LOGS)

    load_settings()
    ligor = [l for l in load_ligor() if l.get("active")]
    log(f"Loaded {len(ligor)} active league(s)")

    data = load_data()
    active_ids = {l["id"] for l in ligor}

    new_games = 0
    for liga in tqdm(ligor, desc="Fetching fixtures", unit="league"):
        league = ensure_league(data, liga)
        name = liga.get("name") or liga.get("country") or "?"
        match_urls = extract_result_links(name, liga.get("url", ""))
        new_count = 0
        for m in match_urls:
            gid = m.rstrip("/").rsplit("/", 1)[-1]
            if gid not in league.get("Games", {}):
                league["Games"][gid] = {"Url": m, "Status": "pending"}
                new_games += 1
                new_count += 1
        if match_urls:
            tag = f"{new_count} new" if new_count else "0 new"
            log(f"[{name}] {len(match_urls)} on page, {tag}", "B")

    save_data(data)

    pending = sum(1 for lg in data["Leagues"].values() for g in lg.get("Games", {}).values() if g.get("Status") == "pending")
    log(f"Total in queue: {pending} pending.", "C")

    success = process_pending_games(data, active_ids)
    save_data(data)

    written = export_csv(data, csv_path, active_ids=active_ids)
    log(f"Final: {success} succeeded. Exported {written} rows.", "BG" if success else "Y")
    log(f"Output: {csv_path}")
    log(f"Data: {DATA_FILE}")
    log(f"Total time: {format_duration((datetime.now() - t_start).total_seconds())}", "M")

    log_handle.close()


if __name__ == "__main__":
    main()
